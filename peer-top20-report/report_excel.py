"""Excel 报告导出"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from data_processor import CategoryGroup, KeywordReport

HEADERS = [
    "排名",
    "公司名称",
    "店铺链接",
    "主营产品",
    "产品总数",
    "类目访问",
    "类目询盘",
    "询盘率",
    "全店线上订单量",
    "全店线上订单金额",
    "商家星等级",
    "供应商年限",
]

SUMMARY_LABELS = [
    "同行平均统计",
    "",
    "",
    "",
    "产品总数",
    "类目访问",
    "类目询盘",
    "询盘率",
    "全店线上订单量",
    "全店线上订单金额",
    "商家星等级",
    "供应商年限",
]


def _style_header(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="FAFAFA")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_border(row_cells) -> None:
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in row_cells:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_category_sheet(sheet, keyword: str, category: CategoryGroup) -> None:
    sheet.title = category.category[:31] or "Top20"
    sheet["A1"] = f"{category.category} 类目询盘同行 top20 排行榜"
    sheet["A1"].font = Font(size=14, bold=True)
    sheet.merge_cells("A1:L1")
    sheet["A2"] = f"关键词：{keyword}"
    sheet.merge_cells("A2:L2")

    header_row = 4
    for col, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=header_row, column=col, value=header)
        _style_header(cell)

    current_row = header_row + 1
    for row in category.rows:
        values = [
            row.rank,
            row.company_name,
            row.home,
            row.main_products,
            row.total_product_count,
            row.page_views,
            row.inquiries,
            row.inquiry_rate,
            row.transaction_number,
            row.transaction_price,
            row.display_star_level,
            row.supplier_year,
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(row=current_row, column=col, value=value)
        _apply_border(sheet[current_row])
        current_row += 1

    summary = category.summary
    summary_values = [
        "同行平均统计",
        "",
        "",
        "",
        summary["total_product_count"],
        summary["page_views"],
        summary["inquiries"],
        summary["inquiry_rate"],
        summary["transaction_number"],
        summary["transaction_price"],
        summary["display_star_level"],
        summary["supplier_year"],
    ]
    for col, value in enumerate(summary_values, start=1):
        cell = sheet.cell(row=current_row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor="FFFBE6")
        cell.font = Font(bold=True)
    _apply_border(sheet[current_row])

    for col in range(1, len(HEADERS) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 16
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 24


def export_excel_report(
    reports: list[KeywordReport],
    output_path: str | Path,
    *,
    selected_category: str | None = None,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for report in reports:
        category_name = selected_category or report.default_category
        category = next((c for c in report.categories if c.category == category_name), None)
        if category is None and report.categories:
            category = report.categories[0]
        if category is None:
            continue

        sheet_name = f"{report.keyword}-{category.category}"[:31]
        sheet = workbook.create_sheet(title=sheet_name)
        _write_category_sheet(sheet, report.keyword, category)

    if not workbook.sheetnames:
        raise ValueError("没有可导出的类目数据")

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
